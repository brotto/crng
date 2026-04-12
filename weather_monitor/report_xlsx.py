#!/usr/bin/env python3
"""
Weather Monitor — XLSX Report Generator
=========================================
Generates Excel reports from the SQLite database.

Usage:
  python report_xlsx.py                          # Today's report
  python report_xlsx.py --date 2026-04-06        # Specific date
  python report_xlsx.py --range 2026-04-01 2026-04-10  # Date range
"""

import argparse
import os
import sys
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(__file__))
from db import get_connection, init_db

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter


REPORT_DIR = os.path.join(os.path.dirname(__file__), "data")

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CRNG_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
OM_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WINNER_FONT = Font(bold=True, color="006100")
LOSER_FONT = Font(color="9C0006")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

CITIES = {
    "sp": "Sao Paulo", "nyc": "New York", "london": "London"
}


def style_header(ws, row, max_col):
    """Apply header styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def generate_report(target_date=None, end_date=None):
    """Generate complete XLSX report."""
    if target_date is None:
        target_date = datetime.now().strftime("%Y-%m-%d")
    if end_date is None:
        end_date = target_date

    conn = get_connection()
    wb = openpyxl.Workbook()

    # ── Sheet 1: WF1 Forecast Tracking ──────────────────────
    ws1 = wb.active
    ws1.title = "WF1 - Forecasts"

    headers = ["City", "Date", "Hour", "Capture Time", "Temp (°C)",
               "Humidity (%)", "Cloud (%)", "Precip (mm)", "P(rain) (%)",
               "Wind (km/h)", "Pressure (hPa)", "Weather"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers))

    rows = conn.execute("""
    SELECT city, target_date, target_hour, captured_at,
           temperature_2m, relative_humidity_2m, cloud_cover,
           precipitation, precipitation_probability,
           wind_speed_10m, pressure_msl, weather_code
    FROM wf1_forecasts
    WHERE target_date BETWEEN ? AND ?
    ORDER BY city, target_date, target_hour, captured_at
    """, (target_date, end_date)).fetchall()

    for i, r in enumerate(rows, 2):
        vals = [CITIES.get(r['city'], r['city']), r['target_date'], r['target_hour'],
                r['captured_at'][:16], r['temperature_2m'], r['relative_humidity_2m'],
                r['cloud_cover'], r['precipitation'], r['precipitation_probability'],
                r['wind_speed_10m'], r['pressure_msl'], r['weather_code']]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER

    # ── Sheet 2: WF1 Drift Analysis ────────────────────────
    ws1d = wb.create_sheet("WF1 - Drift")
    headers = ["City", "Date", "Hour", "First Forecast", "Latest Forecast",
               "Drift (°C)", "N Captures"]
    for col, h in enumerate(headers, 1):
        ws1d.cell(row=1, column=col, value=h)
    style_header(ws1d, 1, len(headers))

    drift_rows = conn.execute("""
    SELECT city, target_date, target_hour,
           MIN(temperature_2m) as t_first,
           MAX(temperature_2m) as t_last,
           MAX(temperature_2m) - MIN(temperature_2m) as drift,
           COUNT(DISTINCT captured_at) as n_captures
    FROM wf1_forecasts
    WHERE target_date BETWEEN ? AND ?
    GROUP BY city, target_date, target_hour
    HAVING n_captures > 1
    ORDER BY drift DESC
    """, (target_date, end_date)).fetchall()

    for i, r in enumerate(drift_rows, 2):
        vals = [CITIES.get(r['city'], r['city']), r['target_date'], r['target_hour'],
                r['t_first'], r['t_last'], r['drift'], r['n_captures']]
        for col, v in enumerate(vals, 1):
            cell = ws1d.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER
            if col == 6 and v and v > 0.5:
                cell.font = Font(bold=True, color="FF0000")

    # ── Sheet 3: WF2 Observations ──────────────────────────
    ws2 = wb.create_sheet("WF2 - Observations")
    headers = ["City", "Date", "Hour", "Temp (°C)", "Humidity (%)",
               "Cloud (%)", "Precip (mm)", "Wind (km/h)", "Weather"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers))

    obs_rows = conn.execute("""
    SELECT city, observed_date, observed_hour,
           temperature_2m, relative_humidity_2m, cloud_cover,
           precipitation, wind_speed_10m, weather_code
    FROM wf2_observations
    WHERE observed_date BETWEEN ? AND ?
    ORDER BY city, observed_date, observed_hour
    """, (target_date, end_date)).fetchall()

    for i, r in enumerate(obs_rows, 2):
        vals = [CITIES.get(r['city'], r['city']), r['observed_date'], r['observed_hour'],
                r['temperature_2m'], r['relative_humidity_2m'], r['cloud_cover'],
                r['precipitation'], r['wind_speed_10m'], r['weather_code']]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER

    # ── Sheet 4: WF3 Predictions ───────────────────────────
    ws3 = wb.create_sheet("WF3 - CRNG Predictions")
    headers = ["City", "Date", "Hour", "Predicted At", "Temp Pred",
               "CI Lo", "CI Hi", "Cloud Factor", "Diurnal dT",
               "Base Temp", "Adjusted?", "Reason"]
    for col, h in enumerate(headers, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers))

    pred_rows = conn.execute("""
    SELECT city, target_date, target_hour, predicted_at,
           temperature_pred, temperature_ci_lo, temperature_ci_hi,
           cloud_factor, diurnal_dt, base_temperature,
           is_adjustment, adjustment_reason
    FROM wf3_predictions
    WHERE target_date BETWEEN ? AND ?
    ORDER BY city, target_date, target_hour, predicted_at
    """, (target_date, end_date)).fetchall()

    for i, r in enumerate(pred_rows, 2):
        vals = [CITIES.get(r['city'], r['city']), r['target_date'], r['target_hour'],
                r['predicted_at'][:16], r['temperature_pred'],
                r['temperature_ci_lo'], r['temperature_ci_hi'],
                r['cloud_factor'], r['diurnal_dt'], r['base_temperature'],
                "Yes" if r['is_adjustment'] else "No", r['adjustment_reason'] or ""]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER
            if col == 11 and v == "Yes":
                cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")

    # ── Sheet 5: Scorecard ─────────────────────────────────
    ws4 = wb.create_sheet("Scorecard")
    headers = ["City", "Date", "Hour", "Observed", "CRNG Pred",
               "OM Forecast", "Err CRNG", "Err OM", "Winner",
               "In CI?", "CI Width"]
    for col, h in enumerate(headers, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers))

    # Build scorecard from joined data
    score_data = conn.execute("""
    SELECT
        o.city, o.observed_date, o.observed_hour, o.temperature_2m as obs_temp,
        p.temperature_pred as crng_temp,
        p.temperature_ci_lo, p.temperature_ci_hi,
        f.temperature_2m as om_temp
    FROM wf2_observations o
    LEFT JOIN (
        SELECT city, target_date, target_hour, temperature_pred,
               temperature_ci_lo, temperature_ci_hi,
               MAX(predicted_at) as latest
        FROM wf3_predictions
        GROUP BY city, target_date, target_hour
    ) p ON o.city = p.city AND o.observed_date = p.target_date AND o.observed_hour = p.target_hour
    LEFT JOIN (
        SELECT city, target_date, target_hour, temperature_2m,
               MIN(captured_at) as first
        FROM wf1_forecasts
        GROUP BY city, target_date, target_hour
    ) f ON o.city = f.city AND o.observed_date = f.target_date AND o.observed_hour = f.target_hour
    WHERE o.observed_date BETWEEN ? AND ?
    AND o.temperature_2m IS NOT NULL
    ORDER BY o.city, o.observed_date, o.observed_hour
    """, (target_date, end_date)).fetchall()

    row_idx = 2
    for r in score_data:
        obs = r['obs_temp']
        crng = r['crng_temp']
        om = r['om_temp']

        err_c = abs(crng - obs) if crng and obs else None
        err_o = abs(om - obs) if om and obs else None
        winner = ""
        if err_c is not None and err_o is not None:
            winner = "CRNG" if err_c < err_o else ("OM" if err_o < err_c else "DRAW")

        in_ci = ""
        ci_width = ""
        if crng and r['temperature_ci_lo'] and r['temperature_ci_hi']:
            in_ci = "Yes" if r['temperature_ci_lo'] <= obs <= r['temperature_ci_hi'] else "No"
            ci_width = round(r['temperature_ci_hi'] - r['temperature_ci_lo'], 1)

        vals = [
            CITIES.get(r['city'], r['city']), r['observed_date'], r['observed_hour'],
            obs, crng, om,
            round(err_c, 1) if err_c else "",
            round(err_o, 1) if err_o else "",
            winner, in_ci, ci_width
        ]
        for col, v in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col, value=v)
            cell.border = THIN_BORDER
            if col == 9:
                if v == "CRNG":
                    cell.font = WINNER_FONT
                    cell.fill = CRNG_FILL
                elif v == "OM":
                    cell.font = LOSER_FONT
                    cell.fill = OM_FILL
        row_idx += 1

    # Auto-fit columns
    for ws in [ws1, ws1d, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 25)

    # Save
    filename = f"weather_report_{target_date}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)
    print(f"\nReport saved: {filepath}")
    conn.close()
    return filepath


if __name__ == "__main__":
    init_db()

    parser = argparse.ArgumentParser(description="Generate XLSX weather report")
    parser.add_argument("--date", help="Target date (YYYY-MM-DD)")
    parser.add_argument("--range", nargs=2, metavar=("START", "END"), help="Date range")
    args = parser.parse_args()

    if args.range:
        generate_report(args.range[0], args.range[1])
    else:
        generate_report(args.date)
